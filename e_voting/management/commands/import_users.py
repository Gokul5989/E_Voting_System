# e_voting/management/commands/import_users.py
import random, string
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.core.mail import send_mail
from e_voting.models import User, UserProfile

def generate_voter_id():
    return ''.join(random.choices(string.digits, k=10))

class Command(BaseCommand):
    help = 'Import users from Excel and generate voter IDs'

    def handle(self, *args, **kwargs):
        wb = load_workbook('data/users.xlsx')
        sheet = wb['users']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username, email, role = row
            if User.objects.filter(username=username).exists():
                continue

            user = User.objects.create_user(username=username, email=email, password='defaultpass123')
            voter_id = generate_voter_id()
            UserProfile.objects.create(user=user, role=role, voter_id=voter_id)

            # Send Voter ID via email
            send_mail(
                subject="Your Voter ID - E-Voting System",
                message=f"Dear {username},\nYour unique 10-digit Voter ID is: {voter_id}\nUse this ID to cast your vote.",
                from_email="admin@evoting.com",
                recipient_list=[email],
                fail_silently=False,
            )
        print("Users imported and emails sent.")
